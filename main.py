import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles import Border, Side
import string
import json
from io import BytesIO

# Tip calculator

# Define the data format for a week of tips, helper methods for creating an initially empty week, and printing a week

data_format_example = {
    'Mo': {
        'metadata': {
            'total_tip': 123,
            'total_hours': 3,
        },
        'employees': {
            'Aaron Eißner': {
                'hours': 3,
                'tip': 0,
            }
        }
    },
    'Di': {

    },
    # ...
    'We': { # WochenEnde
        'metadata': {
            'total_tip': 123,
            'total_hours': 3,
        },
        'employees': {
            'Aaron Eißner': {
                'hours': 3,
                'tip': 0,
            }
        }
    }
}

days = ['Mo', 'Di', 'Mi', 'Do', 'Fr', 'Sa', 'So']
day_titles = ['Montag', 'Dienstag', 'Mittwoch', 'Donnerstag', 'Freitag', 'Samstag', 'Sonntag']

def create_metadata(total_daily_tip):
    return {
        'total_tip': total_daily_tip,
        'total_hours': 0
    }

def create_employees(employee_names):
    return { 
        employee_name: { 
            'hours': 0, 
            'tip': 0 
        } for employee_name in employee_names 
    }

def create_day(total_daily_tip, employee_names):
    return {
        'metadata': create_metadata(total_daily_tip),
        'employees': create_employees(employee_names)
    }

def create_week(total_daily_tips, employee_names):
    return {
        'Mo': create_day(total_daily_tips['Mo'], employee_names),
        'Di': create_day(total_daily_tips['Di'], employee_names),
        'Mi': create_day(total_daily_tips['Mi'], employee_names),
        'Do': create_day(total_daily_tips['Do'], employee_names),
        'Fr': create_day(total_daily_tips['Fr'], employee_names),
        'Sa': create_day(total_daily_tips['Sa'], employee_names),
        'So': create_day(total_daily_tips['So'], employee_names),
        'We': create_day(sum(total_daily_tips.values()), employee_names)
    }

def print_week(week):
    pretty_week = json.dumps(week, sort_keys=False, indent=4)
    print(pretty_week)

# Auxiliary functions

def tip(hours, tip_per_hour):
    return round(hours * tip_per_hour, 2)

def tip_per_hour(tip, hours):
    return tip / hours;

def hours_from_decimal(decimal):
    hours = int(decimal)
    minutes = round((decimal - hours) * 60)
    return f"{hours:02}:{minutes:02}"

def decimal_from_hours(hours):
    hours, minutes = map(int, hours.split(':'))
    return round(hours + (minutes / 60), 2) 

# Read the input excel file and add any new employee names

def get_employee_names(xl):
    employee_names = [
        "Alicia Anderson", "Emilie Appl", "Mina Aryal", "Nadine Aschenbrenner",
        "Omar Massire Balde", "Acelya Basili", "Fabrizio Bellia", "Filippo Bieringer",
        "Anna Sophie Blessing", "Sabullah Bullach", "Manuel Bär", "Nepomuk Böhm",
        "Tom Colombo", "Annika Ebel", "Stina Ebel", "Aaron Eißner", "Jasmin Engelberger",
        "Alicia Ekoos", "Paul Firmhofer", "Markus Fuchs", "Marie Gaschler", "Lisa Gierer",
        "Falk Golla", "Dodo Goßner", "Ekaterina Grashchenko", "Laura Greidenweiß",
        "Karen Gronbach", "Miriam Guggenberger", "Svenja Halamek", "Nicklas Herold",
        "Katrin Hiller", "Marie Jaussi", "Kuda Kabylbekov", "Angela Knoll",
        "Claudine-Sophie Kopetz", "Elina Kuldeva", "Philippe Leonpacher",
        "Moritz Liederscheidt", "Marlene Lind", "Anton Machowski", "Marleen Mulzer",
        "Simon Männlein", "Sarah-Michelle Müller", "Jannik Mülhaupt", 
        "Okechulowu (Anselem) Ohaebism", "Philine Ostermayer", "Damiano Parziale",
        "Alyah Pattis", "Luca Principi", "Reshimi Raj Aryal", "Lilian Reiner",
        "Emilia Sachsen-Coburg", "Luca Sawallisch", "Pia Schmolly", "Kilian Schugsties",
        "Seyed Mohammad", "Mauro Sirigu", "Matthias Strobel", "Tanja Stürhof",
        "Luca Tegeder", "Fenny Tran", "Bent van Zon", "Charlotte von Schröder",
        "Jana Wegenke", "Lilli Wittig", "Jennifer Wöhrlin", "Fabian Zink", "Anna Graf",
        "Emelie Appl", "Açelya Basili", "Sabulla Bullach", "Okechulowu Ohaebsim", "Luca Principi co Bischoff"
    ]

    weekly_employee_names = xl.sheet_names
    weekly_employee_names.remove('Übersicht')

    employee_names = sorted(set(employee_names + weekly_employee_names))
    employee_names.sort()

    return weekly_employee_names, employee_names

# Create an initially empty week of tips
# Iterate through the input excel file sheet by sheet and set the total_hours for each day and hours for each day and each employee

def parse_input_excel_for_employee(xl, employee_name, week):
    sheet = xl.parse(employee_name, header=None)
    header = sheet[sheet[0] == 'Tag'].index[0]

    data = xl.parse(employee_name, header=header)
    data = data[['Tag', 'Startzeit', 'Endzeit', 'Pause (min)', 'Dauer netto (h)']]
    data = data[data['Startzeit'].notnull()]
    
    for _, row in data.iterrows():
        day = row['Tag'][0:2]
        hours = decimal_from_hours(row['Dauer netto (h)'])
        week[day]['metadata']['total_hours'] += hours
        week[day]['employees'][employee_name]['hours'] += hours

def parse_input_excel_for_employees(xl, weekly_employee_names, week):
    for employee_name in weekly_employee_names:
        parse_input_excel_for_employee(xl, employee_name, week)

# Iterate through the days and calculate the tip for each day for each employee

def calculate_tips(week):
    def go(week, day_name):
        day = week[day_name]
        metadata = day['metadata']

        tph = tip_per_hour(metadata['total_tip'], metadata['total_hours'])

        for employee in day['employees'].values():
            employee['tip'] = tip(employee['hours'], tph)

    for day in days:
        go(week, day)

# Iterate through each day and calculate the summary for the whole week

def calculate_summary(week):
    we = week['We']

    for day in days:
        we['metadata']['total_hours'] += week[day]['metadata']['total_hours']
        for employee_name, hours_tip in week[day]['employees'].items():
            we['employees'][employee_name]['hours'] += hours_tip['hours']
            we['employees'][employee_name]['tip'] += hours_tip['tip']

    for employee_name, hours_tip in week[day]['employees'].items():
        we['employees'][employee_name]['hours'] = round(we['employees'][employee_name]['hours'], 2)
        we['employees'][employee_name]['tip'] = round(we['employees'][employee_name]['tip'], 2)

def calculate_tip_week(xl):
    weekly_employee_names, employee_names = get_employee_names(xl)
    week = create_week(total_daily_tips, employee_names)
    parse_input_excel_for_employees(xl, weekly_employee_names, week)
    calculate_tips(week)
    calculate_summary(week)

    return week, employee_names, weekly_employee_names

# Create an output excel file from the weekly tips

def calculate_output_excel(week, employee_names):
    wb = Workbook()
    ws = wb.active

    cols = list(string.ascii_uppercase) + [f'{letter}{next_letter}' for letter in string.ascii_uppercase for next_letter in string.ascii_uppercase[:1]]
    column_width = 20
    employee_start_row = 6

    def set_column_width(col_width, cols):
        for col in cols:
            ws.column_dimensions[col].width = col_width

    def set_borders(until_row, cols):
        border_style_right = Border(
            right=Side(border_style='thin', color='000000')
        )

        for row in range(1, until_row):
            for col in cols:
                ws[f'{col}{row}'].border = border_style_right

        border_style_all = Border(
            top=Side(border_style='thin', color='000000'),
            right=Side(border_style='thin', color='000000'),
            bottom=Side(border_style='thin', color='000000'),
            left=Side(border_style='thin', color='000000'),
        )

        for col in cols:
            ws[f'{col}5'].border = border_style_all

        border_style_top = Border(
            top=Side(border_style='thin', color='000000')
        )

        for col in cols:
            ws[f'{col}{until_row}'].border = border_style_top

    def write_employee_col(employee_names):
        ws['A5'] = 'Mitarbeiter'
        ws['A5'].font = Font(bold=True)

        for row, employee_name in enumerate(employee_names, start=employee_start_row):
            ws[f'A{row}'] = employee_name

    def write_day_header(col, day, total_tip, total_hours, total_tip_per_hour):
        col_index = cols.index(col)
        col_1 = cols[col_index]
        col_2 = cols[col_index+1]
        col_3 = cols[col_index+2]

        ws.merge_cells(f'{col_1}1:{col_3}1')
        ws[f'{col_1}1'] = day
        ws[f'{col_1}1'].font = Font(size=13, bold=True)
        ws[f'{col_1}1'].alignment = Alignment(horizontal='center', vertical='center')

        ws.merge_cells(f'{col_1}2:{col_2}2')
        ws[f'{col_1}2'] = 'Trinkgeld (gesamt)'
        ws[f'{col_3}2'] = total_tip
        ws[f'{col_3}2'].number_format = '"€"#,##0.00'

        ws.merge_cells(f'{col_1}3:{col_2}3')
        ws[f'{col_1}3'] = 'Stunden (gesamt)'
        ws[f'{col_3}3'] = total_hours
        ws[f'{col_3}3'].alignment = Alignment(horizontal='right')

        ws.merge_cells(f'{col_1}4:{col_2}4')
        ws[f'{col_1}4'] = 'Trinkgeld pro Stunde'
        ws[f'{col_3}4'] = total_tip_per_hour
        ws[f'{col_3}4'].alignment = Alignment(horizontal='right')

        ws[f'{col_1}5'] = 'Stunden'
        ws[f'{col_1}5'].font = Font(bold=True)

        ws[f'{col_2}5'] = 'Stunden (Dezimal)'
        ws[f'{col_2}5'].font = Font(bold=True)

        ws[f'{col_3}5'] = 'Trinkgeld'
        ws[f'{col_3}5'].font = Font(bold=True)

    def write_day_row(col, row, hours, hours_decimal, tip):
        if (hours_decimal == 0):
            return

        col_index = cols.index(col)
        col_1 = cols[col_index]
        col_2 = cols[col_index+1]
        col_3 = cols[col_index+2]

        ws[f'{col_1}{row}'] = hours
        ws[f'{col_1}{row}'].alignment = Alignment(horizontal='right')
        ws[f'{col_2}{row}'] = hours_decimal
        ws[f'{col_3}{row}'] = tip
        ws[f'{col_3}{row}'].number_format = '"€"#,##0.00'

    def write_day_col(day_key, day_title, col, week, employee_names):
        day = week[day_key]
        metadata = day['metadata']
        employees = day['employees']

        write_day_header(col, day_title, metadata['total_tip'], metadata['total_hours'], tip_per_hour(metadata['total_tip'], metadata['total_hours']))

        for row, employee_name in enumerate(employee_names, start=6):
            employee = employees[employee_name]
            hours_decimal = employee['hours']
            hours = hours_from_decimal(hours_decimal)
            tip = employee['tip']
            write_day_row(col, row, hours, hours_decimal, tip)

    def write_week(week, employee_names):
        for day_key, day_title, col in zip(days + ['We'], day_titles + ['Woche'], ['B', 'E', 'H', 'K', 'N', 'Q', 'T', 'W']):
            write_day_col(day_key, day_title, col, week, employee_names)

    set_column_width(column_width, cols[:cols.index('Z')])
    set_borders(len(employee_names) + 6, cols[:cols.index('Z')])
    write_employee_col(employee_names)
    write_week(week, employee_names)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output

# Streamlit

total_daily_tips = {}

st.title('Trinkgeld Rechner')

st.write('Tägliches Trinkgeld')

columns = st.columns(len(days))
for col, day, day_title in zip(columns, days, day_titles):
    total_daily_tips[day] = col.number_input(f'{day_title}:', min_value=0.0, value=0.0, step=0.01, format='%.2f')

st.write('Tägliche Arbeitszeiten')
uploaded_file = st.file_uploader('Bitte Excel Datei hochladen', type=['xls', 'xlsx'])

if uploaded_file is not None:
    xl = pd.ExcelFile(uploaded_file)

    st.write(f'Trinkgeld wurde berechnet')

    week, employee_names, weekly_employee_names = calculate_tip_week(xl)
    tip_excel = calculate_output_excel(week, employee_names)

    st.download_button('Download', tip_excel, 'trinkgeld.xlsx')


#new 
